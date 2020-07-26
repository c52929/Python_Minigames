import xlrd
import openpyxl

words=xlrd.open_workbook('words_data.xlsx').sheet_by_index(0).col_values(0)
# for i in range(10):
# 	print(words[i])
# print()

wb=openpyxl.load_workbook('words_data.xlsx')
# # wbcopy=openpyxl.load_workbook('words_data-copy.xlsx')
# print(type(wb))
# print(wb.sheetnames)
# # print(type(wbcopy))
wsn=wb.create_sheet(title="after")
# print(wb.sheetnames)
# # print(wbcopy.sheetnames)

# 臨時
# words=['sup','hey!','aaaaaaaap','yaaaaaaay!','pen','no!','going']
# print(words)

deleted=[]

for i in range(len(words)):
	delete=0
	letters=[]
	for j in range(len(words[i])):
		if not(words[i][j] in letters):
			if words[i][j].upper() in['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V','W','X','Y','Z']:
				letters.append(words[i][j])
			else:
				delete=1
	if len(letters)<=2 or delete>0:
		deleted.append(words[i])
		i-=1
for i in range(len(deleted)):
	del words[words.index(deleted[i])]

# print(wsn.cell(row=1, column=1).value)
# print()

# print(wsn["A1:A3"])
# print(wsn["A1:A3"])

for i in range(len(words)):
	wsn.cell(row=i+1, column=1, value=words[i])

# for i in range(len(words)):
# 	print(wsn.cell(row=i+1, column=1).value)

# print(wsn.cell(row=1, column=1).value)
# print(wsn.cell(row=2, column=1).value)
# print(wsn.cell(row=3, column=1).value)
# print(words)
print(deleted)

wb.save('words_data.xlsx')